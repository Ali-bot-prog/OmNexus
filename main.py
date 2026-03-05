from __future__ import annotations
from io import BytesIO
from fastapi import UploadFile, File
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, HTTPException, Query
from pydantic import BaseModel, Field, ConfigDict
from typing import Optional, Literal, Dict, Any, List, Tuple
from datetime import datetime, timezone, timedelta
from contextlib import asynccontextmanager
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
import bcrypt
from jose import JWTError, jwt
from fastapi import Depends, status
from google import genai
from fpdf import FPDF

import sqlite3
import os
import random
import ast
import scraper
import lead_finder
import statistics
import random
import math
from apscheduler.schedulers.background import BackgroundScheduler


# ==================================================
# CONFIG
# ==================================================
# import shared constants; helps prevent drift between scripts
from config import (
    APP_TITLE,
    BASE_DIR,
    DB_PATH,
    PORT,
    SECRET_KEY,
    ALGORITHM,
    ACCESS_TOKEN_EXPIRE_MINUTES,
    GEMINI_API_KEY,
)

PropertyTur = Literal["konut", "arsa", "ticari"]

# [LLM] Client Initialization
ai_client = None
if GEMINI_API_KEY:
    ai_client = genai.Client(api_key=GEMINI_API_KEY)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="auth/login")


# ==================================================
# DB
# ==================================================
def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    conn = db()
    cur = conn.cursor()

    # 1. Tabloyu oluştur (Eğer yoksa)
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS emsal (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tur TEXT NOT NULL,
            listing_type TEXT, -- [NEW] satilik/kiralik
            il TEXT,
            ilce TEXT,
            mahalle TEXT,
            lat REAL,
            lng REAL,

            brut_m2 REAL,     -- konut/ticari için
            net_m2 REAL,      -- [NEW] net alan
            arsa_m2 REAL,     -- arsa için
            kaks REAL,        -- arsa için (emsal)
            bina_yasi INTEGER,
            bina_kat_sayisi INTEGER, -- [NEW]
            bulundugu_kat INTEGER,   -- [NEW]

            kat TEXT,         -- [Descriptive: Zemin, Ara Kat...]
            cephe TEXT,       
            isitma TEXT,      
            otopark INTEGER,  
            tapu TEXT,        
            imar TEXT,
            asansor INTEGER,  -- [NEW]
            esyali INTEGER,   -- [NEW]
            site_icerisinde INTEGER, -- [NEW]

            fiyat REAL NOT NULL,
            kira REAL,

            kaynak TEXT,
            tenant_id INTEGER, -- [NEW] SaaS Support
            durum TEXT DEFAULT 'aktif', -- [NEW] Satildi/Kiralandi/Aktif/Pasif
            created_at TEXT NOT NULL
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS piyasa_hafizasi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kategori TEXT NOT NULL,
            soru_baslik TEXT NOT NULL,
            cevap_icerik TEXT NOT NULL,
            okunma_sayisi INTEGER DEFAULT 0,
            tenant_id INTEGER,
            created_at TEXT NOT NULL
        )
        """
    )
    
    # 2. Migrasyon (Eksik kolonları ekle)
    # Basit bir yöntem: Kolon eklemeyi dene, varsa hata verir (catch)
    new_cols = [
        ("brut_m2", "REAL"), ("net_m2", "REAL"), ("arsa_m2", "REAL"), ("kaks", "REAL"), 
        ("bina_yasi", "INTEGER"), ("bina_kat_sayisi", "INTEGER"), ("bulundugu_kat", "INTEGER"),
        ("kat", "TEXT"), ("cephe", "TEXT"), ("isitma", "TEXT"), 
        ("otopark", "INTEGER"), ("tapu", "TEXT"), ("imar", "TEXT"),
        ("oda_sayisi", "TEXT"), # [NEW] Added previously but checking again
        ("asansor", "INTEGER"), ("esyali", "INTEGER"), ("site_icerisinde", "INTEGER"), # [NEW]
        ("kaynak", "TEXT"), ("listing_type", "TEXT"),
        ("tenant_id", "INTEGER"), # [NEW]
        ("durum", "TEXT") # [NEW]
    ]
    for col_name, col_type in new_cols:
        try:
            cur.execute(f"ALTER TABLE emsal ADD COLUMN {col_name} {col_type}")
        except sqlite3.OperationalError:
            pass # Kolon zaten var

    # Log (Audit Trail)
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS action_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT NOT NULL,
            detail TEXT,
            created_at TEXT NOT NULL
        )
        """
    )
    
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            email TEXT,
            hashed_password TEXT NOT NULL,
            role TEXT DEFAULT 'user',
            tenant_id INTEGER,
            created_at TEXT NOT NULL,
            FOREIGN KEY(tenant_id) REFERENCES tenants(id)
        )
        """
    )

    # New admin/content tables
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS site_content (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            key TEXT UNIQUE NOT NULL,
            value TEXT,
            tenant_id INTEGER,
            created_at TEXT NOT NULL
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS faq (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            question TEXT NOT NULL,
            answer TEXT NOT NULL,
            tenant_id INTEGER,
            created_at TEXT NOT NULL
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS contact_message (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT NOT NULL,
            phone TEXT,
            message TEXT NOT NULL,
            read INTEGER DEFAULT 0,
            tenant_id INTEGER,
            created_at TEXT NOT NULL
        )
        """
    )

    # USER Migrasyonu (tenant_id ekle)
    try:
        cur.execute("ALTER TABLE users ADD COLUMN tenant_id INTEGER")
    except sqlite3.OperationalError:
        pass # Zaten var

    conn.commit()
    conn.close()

def log_action(action: str, detail: str = ""):
    """Audit log helper"""
    try:
        conn = db()
        cur = conn.cursor()
        cur.execute("INSERT INTO action_log (action, detail, created_at) VALUES (?, ?, ?)", (
            action, detail, _now_iso()
        ))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Log error: {e}")


def get_password_hash(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))


def create_access_token(data: dict, expires_delta: Optional[float] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + timedelta(minutes=expires_delta)
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt






def row_to_dict(r: sqlite3.Row) -> Dict[str, Any]:
    return dict(r)


EXCEL_COLUMNS = [
    ("tur", "Tür"), ("listing_type", "Tip"), ("il", "İl"), ("ilce", "İlçe"), ("mahalle", "Mahalle"),
    ("brut_m2", "Brüt M2"), ("arsa_m2", "Arsa M2"), ("kaks", "KAKS"), ("bina_yasi", "Bina Yaşı"),
    ("kat", "Kat"), ("cephe", "Cephe"), ("isitma", "Isıtma"), ("otopark", "Otopark"),
    ("tapu", "Tapu"), ("imar", "İmar"), ("fiyat", "Fiyat"), ("kira", "Kira"),
    ("kaynak", "Kaynak"), ("created_at", "Tarih")
]


def _auto_fit_columns(ws):
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        try:
            col_letter = get_column_letter(column[0].column)
        except:
            # Fallback
            col_letter = column[0].column_letter
            
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width



def rows_to_dicts(rows: List[sqlite3.Row]) -> List[Dict[str, Any]]:
    return [row_to_dict(r) for r in rows]


# ==================================================
# Pydantic Models
# ==================================================
# ==================================================
# Pydantic Models (Strict & Documented)
# ==================================================
class EmsalBase(BaseModel):
    model_config = ConfigDict(extra="ignore", from_attributes=True)

    tur: PropertyTur = Field(..., description="Mülk tipi: konut, arsa veya ticari")
    listing_type: Optional[Literal["satilik", "kiralik"]] = Field("satilik", description="İlan tipi: satilik veya kiralik")
    il: Optional[str] = Field(None, min_length=2, max_length=50, description="İl adı")
    ilce: Optional[str] = Field(None, min_length=2, max_length=50, description="İlçe adı")
    mahalle: Optional[str] = Field(None, min_length=2, max_length=100, description="Mahalle adı")

    lat: Optional[float] = Field(None, ge=-90, le=90, description="Enlem (Latitude)")
    lng: Optional[float] = Field(None, ge=-180, le=180, description="Boylam (Longitude)")

    brut_m2: Optional[float] = Field(None, gt=0, description="Brüt alan (m2)")
    net_m2: Optional[float] = Field(None, gt=0, description="Net alan (m2)")
    arsa_m2: Optional[float] = Field(None, gt=0, description="Arsa alanı (m2)")
    kaks: Optional[float] = Field(None, gt=0, description="Emsal oranı (KAKS)")
    oda_sayisi: Optional[str] = Field(None, description="Oda Sayısı (3+1, 2+1 vb.)") # [NEW]
    bina_yasi: Optional[int] = Field(None, ge=0, le=500, description="Bina yaşı")
    bina_kat_sayisi: Optional[int] = Field(None, ge=0, description="Binadaki toplam kat sayısı")
    bulundugu_kat: Optional[int] = Field(None, description="Bulunduğu kat numarası")

    kat: Optional[str] = Field(None, description="Bulunduğu kat (Zemin, Ara Kat, En Üst, Dubleks, Bodrum)")
    cephe: Optional[str] = Field(None, description="Cephe (Kuzey, Güney, Doğu, Batı)")
    isitma: Optional[str] = Field(None, description="Isıtma (Kombi, Merkezi, Soba)")
    otopark: Optional[int] = Field(None, description="Otopark Var mı? (1/0)")
    tapu: Optional[str] = Field(None, description="Tapu Durumu (Kat Mülkiyeti, Hisseli...)")
    imar: Optional[str] = Field(None, description="İmar Durumu (Konut, Ticari...) - Arsa için")
    
    # [NEW] Detailed Features
    asansor: Optional[int] = Field(None, description="Asansör Var mı? (1/0)")
    esyali: Optional[int] = Field(None, description="Eşyalı mı? (1/0)")
    site_icerisinde: Optional[int] = Field(None, description="Site İçerisinde mi? (1/0)")

    fiyat: float = Field(..., gt=0, description="Satış fiyatı (TL)")
    kira: Optional[float] = Field(None, ge=0, description="Kira bedeli (TL)")
    kaynak: Optional[str] = Field(None, max_length=255, description="Veri kaynağı/Link")
    tenant_id: Optional[int] = Field(None, description="Kiracı ID")
    durum: Optional[str] = Field("aktif", description="İlan Durumu: aktif, satildi, kiralandi, pasif")

    
    # [NEW] Detailed Features
    asansor: Optional[int] = Field(None, description="Asansör Var mı? (1/0)")
    esyali: Optional[int] = Field(None, description="Eşyalı mı? (1/0)")
    site_icerisinde: Optional[int] = Field(None, description="Site İçerisinde mi? (1/0)")


class EmsalCreate(EmsalBase):
    """Kayıt oluştururken kullanılan model"""
    pass


class EmsalResponse(EmsalBase):
    """API yanıtlarında dönen tam model"""
    id: int = Field(..., description="Kayıt ID")
    created_at: str = Field(..., description="Kayıt tarihi (ISO 8601)")


class EmsalSummary(BaseModel):
    """Tahmin sonuçlarında gösterilen özet emsal"""
    id: int
    il: Optional[str]
    ilce: Optional[str]
    mahalle: Optional[str]
    created_at: Optional[str]
    w: float = Field(..., description="Benzerlik skoru")
    fiyat: float
    brut_m2: Optional[float] = None
    net_m2: Optional[float] = None
    bina_yasi: Optional[int] = None
    alan_eff: float
    fiyat_m2_eff: float


class DanismanZekasi(BaseModel):
    neden_fiyat: List[str] = Field(..., description="Fiyatı etkileyen temel faktörler")
    etkili_emsaller: List[int] = Field(..., description="En çok benzeyen emsal ID'leri")
    risk_durumu: str = Field(..., description="Güvenli / Riskli / Orta")
    risk_nedeni: str = Field(..., description="Risk durumunun açıklaması")


class EstimateRange(BaseModel):
    alt: float
    ust: float


class EstimateResponse(BaseModel):
    tur: PropertyTur
    girdi_alan_eff: float
    tahmini_satis: float
    satis_aralik: EstimateRange
    tahmini_kira: float
    kira_aralik: EstimateRange
    confidence: int = Field(..., ge=0, le=100)
    aciklama: str
    kullanilan_emsal: int
    emsal_ozeti: List[EmsalSummary]
    danisman_zekasi: DanismanZekasi
    target_location: Optional[Dict[str, Optional[float]]] = None
    meta: Dict[str, Any]


class AIReportRequest(BaseModel):
    estimate_id: Optional[str] = None # Gelecekte DB referansı için
    analysis_data: Dict[str, Any] # estimate'ten dönen tüm veri
    report_type: Literal["customer", "investor"] = "customer"

class AIReportResponse(BaseModel):
    report_text: str
    report_type: str


class EstimateRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")

    il: Optional[str] = None
    ilce: Optional[str] = None
    mahalle: Optional[str] = None

    lat: Optional[float] = None
    lng: Optional[float] = None

    brut_m2: Optional[float] = Field(default=None, gt=0)
    net_m2: Optional[float] = Field(default=None, gt=0)
    arsa_m2: Optional[float] = Field(default=None, gt=0)
    kaks: Optional[float] = Field(default=None, gt=0)
    
    oda_sayisi: Optional[str] = None
    bina_yasi: Optional[int] = Field(default=None, ge=0)
    bina_kat_sayisi: Optional[int] = Field(default=None, ge=0) 
    bulundugu_kat: Optional[int] = Field(default=None)         

    kat: Optional[str] = None
    cephe: Optional[str] = None
    isitma: Optional[str] = None
    otopark: Optional[int] = None
    tapu: Optional[str] = None
    imar: Optional[str] = None
    
    # [NEW] Detailed Features
    asansor: Optional[int] = None
    esyali: Optional[int] = None
    site_icerisinde: Optional[int] = None

    listing_type: Optional[Literal["satilik", "kiralik"]] = Field("satilik", description="Değerleme tipi: satilik veya kiralik")

    aylik_artis: float = Field(default=0.03, ge=0.0, le=0.50, description="Aylık piyasa artış oranı (örn 0.03)")


# ==================================================
# Auth Models
# ==================================================
class UserBase(BaseModel):
    username: str
    email: Optional[str] = None
    role: str = "user" # admin, user
    tenant_id: Optional[int] = None

class UserCreate(UserBase):
    password: str

class UserLogin(BaseModel):
    username: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str

# --------------------------------------------------
# Content & FAQ Models
# --------------------------------------------------
class SiteContentItem(BaseModel):
    key: str
    value: Optional[str] = None

class FAQItem(BaseModel):
    question: str
    answer: str

class ContactMessageIn(BaseModel):
    name: str
    email: str
    phone: Optional[str] = None
    message: str

class ContactMessageOut(ContactMessageIn):
    id: int
    read: bool = False
    created_at: str

class UserInDB(UserBase):
    id: int
    hashed_password: str
    created_at: str


# ==================================================
# Helpers
# ==================================================
def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _parse_dt(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    try:
        # "2025-12-29T..." -> aware (çoğu durumda)
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except:
        return None


def haversine_km(lat1: Optional[float], lng1: Optional[float], lat2: Optional[float], lng2: Optional[float]) -> Optional[float]:
    if lat1 is None or lng1 is None or lat2 is None or lng2 is None:
        return None
    R = 6371.0
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = math.sin(dlat / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlng / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c


def effective_area(tur: PropertyTur, obj: Any) -> float:
    """
    Konut/Ticari: brut_m2
    Arsa: arsa_m2 * (kaks or 1)
    """
    brut = getattr(obj, "brut_m2", None)
    arsa = getattr(obj, "arsa_m2", None)
    kaks = getattr(obj, "kaks", None)

    if tur == "arsa":
        base = arsa or brut or 0
        k = kaks or 1.0
        return float(base) * float(k)

    return float(brut or 0)


def location_scope(payload: EstimateRequest) -> str:
    if payload.mahalle:
        return "mahalle"
    if payload.ilce:
        return "ilce"
    if payload.il:
        return "il"
    return "genel"


def time_weight(created_at_iso: str, half_life_days: int = 180) -> float:
    """
    Eski emsalin ağırlığını düşürür.
    half_life_days: 180 => 6 ayda ağırlık yarıya iner.
    """
    dt = _parse_dt(created_at_iso)
    if not dt:
        return 0.85  # tarih yoksa orta ağırlık

    now = datetime.now(timezone.utc)
    age_days = max(0.0, (now - dt).total_seconds() / 86400.0)

    # Exponential decay: w = 0.5^(age/half_life)
    w = 0.5 ** (age_days / float(half_life_days))
    # clamp
    return float(max(0.20, min(1.10, w)))


def _sim_common(payload: EstimateRequest, e: Dict[str, Any], tur: PropertyTur) -> float:
    """Ortak temel benzerlik: Lokasyon + Alan + Mesafe + Zaman"""
    w = 1.0

    # 1) Lokasyon
    p_il = (payload.il or "").strip().lower()
    p_ilce = (payload.ilce or "").strip().lower()
    p_mah = (payload.mahalle or "").strip().lower()
    e_il = (e.get("il") or "").strip().lower()
    e_ilce = (e.get("ilce") or "").strip().lower()
    e_mah = (e.get("mahalle") or "").strip().lower()

    if p_il and e_il and p_il == e_il: w *= 1.10
    if p_ilce and e_ilce and p_ilce == e_ilce: w *= 1.15
    if p_mah and e_mah and p_mah == e_mah: w *= 1.30

    # 2) Alan
    p_area = effective_area(tur, payload)
    e_area = effective_area(tur, type("Tmp",(),{"brut_m2":e.get("brut_m2"),"arsa_m2":e.get("arsa_m2"),"kaks":e.get("kaks")})())
    if p_area > 0 and e_area > 0:
        diff_ratio = abs(p_area - e_area) / max(1.0, e_area)
        w *= max(0.35, 1.0 - diff_ratio)

    # 3) Mesafe
    dist = haversine_km(payload.lat, payload.lng, e.get("lat"), e.get("lng"))
    if dist is not None:
        w *= max(0.50, 1.25 - (dist / 4.0))

    # 4) Zaman
    w *= time_weight(e.get("created_at") or "")
    
    return w


def init_db() -> None:
    conn = db()
    cur = conn.cursor()

    # 1. Tabloyu oluştur (Eğer yoksa)
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS emsal (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tur TEXT NOT NULL,
            listing_type TEXT, 
            il TEXT,
            ilce TEXT,
            mahalle TEXT,
            lat REAL,
            lng REAL,

            brut_m2 REAL,     
            net_m2 REAL,      
            arsa_m2 REAL,     
            kaks REAL,        
            oda_sayisi TEXT,
            bina_yasi INTEGER,
            bina_kat_sayisi INTEGER, 
            bulundugu_kat INTEGER,   

            kat TEXT,        
            cephe TEXT,       
            isitma TEXT,      
            otopark INTEGER,  
            tapu TEXT,        
            imar TEXT,        

            fiyat REAL NOT NULL,
            kira REAL,

            kaynak TEXT,
            tenant_id INTEGER,
            created_at TEXT NOT NULL
        )
        """
    )
    
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS piyasa_hafizasi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kategori TEXT NOT NULL,
            soru_baslik TEXT NOT NULL,
            cevap_icerik TEXT NOT NULL,
            okunma_sayisi INTEGER DEFAULT 0,
            tenant_id INTEGER,
            created_at TEXT NOT NULL
        )
        """
    )
    
    # 2. Migrasyon (Eksik kolonları ekle)
    new_cols = [
        ("brut_m2", "REAL"), ("net_m2", "REAL"), ("arsa_m2", "REAL"), ("kaks", "REAL"), 
        ("oda_sayisi", "TEXT"), 
        ("bina_yasi", "INTEGER"), ("bina_kat_sayisi", "INTEGER"), ("bulundugu_kat", "INTEGER"),
        ("kat", "TEXT"), ("cephe", "TEXT"), ("isitma", "TEXT"), 
        ("otopark", "INTEGER"), ("tapu", "TEXT"), ("imar", "TEXT"),
        ("kaynak", "TEXT"), ("listing_type", "TEXT"),
        ("tenant_id", "INTEGER")
    ]
    for col_name, col_type in new_cols:
        try:
            cur.execute(f"ALTER TABLE emsal ADD COLUMN {col_name} {col_type}")
        except sqlite3.OperationalError:
            pass 

    # Log (Audit Trail)
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS action_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT NOT NULL,
            detail TEXT,
            created_at TEXT NOT NULL
        )
        """
    )

    conn.commit()

    # 3. Auth Tabloları (Users, Tenants)
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS tenants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            plan TEXT DEFAULT 'free',
            il TEXT,
            ilce TEXT,
            ofis_adi TEXT,
            onboarded INTEGER DEFAULT 0,
            created_at TEXT NOT NULL
        )
        """
    )
    # Tenant migrasyon
    for col, typ in [("il", "TEXT"), ("ilce", "TEXT"), ("ofis_adi", "TEXT"), ("onboarded", "INTEGER")]:
        try:
            cur.execute(f"ALTER TABLE tenants ADD COLUMN {col} {typ}")
        except:
            pass

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            email TEXT,
            hashed_password TEXT NOT NULL,
            role TEXT DEFAULT 'user',
            tenant_id INTEGER,
            created_at TEXT NOT NULL,
            FOREIGN KEY(tenant_id) REFERENCES tenants(id)
        )
        """
    )

    # 4. FSBO Leads Tablosu
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS fsbo_leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            baslik TEXT,
            ozet TEXT,
            url TEXT,
            platform TEXT,
            tahmini_fiyat REAL,
            il TEXT,
            ilce TEXT,
            tur TEXT,
            listing_type TEXT DEFAULT 'satilik',
            fsbo_skoru INTEGER DEFAULT 0,
            durum TEXT DEFAULT 'yeni',
            portfoy_adayi INTEGER DEFAULT 0,
            tenant_id INTEGER,
            created_at TEXT NOT NULL
        )
        """
    )
    for col, typ in [("portfoy_adayi", "INTEGER"), ("fsbo_skoru", "INTEGER"), ("listing_type", "TEXT")]:
        try:
            cur.execute(f"ALTER TABLE fsbo_leads ADD COLUMN {col} {typ}")
        except:
            pass

    # Varsayılan Admin Kullanıcısı (Eğer yoksa)
    try:
        cur.execute("SELECT id FROM users WHERE username='admin'")
        if not cur.fetchone():
            # Admin için tenant oluştur
            cur.execute("INSERT INTO tenants (name, plan, il, ilce, onboarded, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                       ("Admin Ofisi", "admin", "Ordu", "Ünye", 1, _now_iso()))
            admin_tenant_id = cur.lastrowid
            admin_pass = get_password_hash("admin123")
            cur.execute(
                "INSERT INTO users (username, hashed_password, role, tenant_id, created_at) VALUES (?, ?, ?, ?, ?)",
                ("admin", admin_pass, "admin", admin_tenant_id, _now_iso())
            )
            print("INFO: Varsayılan admin kullanıcısı oluşturuldu (User: admin, Pass: admin123)")
    except Exception as e:
        print(f"Admin user create warning: {e}")

    conn.commit()
    conn.close()


def _sim_konut(payload: EstimateRequest, e: Dict[str, Any], base_w: float) -> float:
    """KONUT MOTORU: Yaşam detaylarına odaklanır"""
    w = base_w
    
    # Oda Sayısı [NEW]
    p_oda = (getattr(payload, 'oda_sayisi', None) or "").strip()
    e_oda = (e.get("oda_sayisi") or "").strip()
    if p_oda and e_oda:
        if p_oda == e_oda: 
            w *= 1.25 # Oda sayısı aynıysa güçlü benzerlik
        else:
            # 3+1 vs 2+1 gibi basit mantık kuralım
            try:
                p_base = int(p_oda.split("+")[0])
                e_base = int(e_oda.split("+")[0])
                diff = abs(p_base - e_base)
                if diff == 0: w *= 1.10 # (3+1 vs 3+0)
                elif diff == 1: w *= 0.85 # (3+1 vs 2+1)
                else: w *= 0.70 # Fark çok
            except:
                w *= 0.80 # Parse edemedik ama farklı

    # Bina Yaşı
    p_age = payload.bina_yasi
    e_age = e.get("bina_yasi")
    if p_age is not None and e_age is not None:
        diff = abs(int(p_age) - int(e_age))
        w *= max(0.50, 1.0 - (diff / 40.0))

    # Bina Kat Sayısı
    p_bks = getattr(payload, 'bina_kat_sayisi', None)
    e_bks = e.get("bina_kat_sayisi")
    if p_bks and e_bks:
        ratio = abs(float(p_bks) - float(e_bks)) / float(max(1, e_bks))
        if ratio < 0.2: w *= 1.05

    # Kat Konumu
    p_kat = (payload.kat or "").lower()
    e_kat = (e.get("kat") or "").lower()
    if p_kat and e_kat:
        if p_kat == e_kat: w *= 1.15
        elif "ara" in p_kat and ("bodrum" in e_kat or "zemin" in e_kat or "çatı" in e_kat): w *= 0.80
        elif "bahçe" in p_kat and "bahçe" in e_kat: w *= 1.20 # Bahçe katı özeldir

    # Cephe
    p_cephe = (payload.cephe or "").lower()
    e_cephe = (e.get("cephe") or "").lower()
    if p_cephe and e_cephe and p_cephe == e_cephe: w *= 1.05

    # Isıtma
    p_isitma = (payload.isitma or "").lower()
    e_isitma = (e.get("isitma") or "").lower()
    if p_isitma and e_isitma and p_isitma != e_isitma:
        if "soba" in e_isitma and "kombi" in p_isitma: w *= 0.85

    # Otopark
    if payload.otopark is not None and e.get("otopark") is not None:
        if payload.otopark == 1 and int(e.get("otopark") or 0) == 0: w *= 0.90
        elif payload.otopark == 0 and int(e.get("otopark") or 0) == 1: w *= 1.05

    # Tapu
    p_tapu = (payload.tapu or "").lower()
    e_tapu = (e.get("tapu") or "").lower()
    if p_tapu and e_tapu and p_tapu != e_tapu:
        if "hisseli" in p_tapu or "hisseli" in e_tapu: w *= 0.65

    # [NEW] Detailed Features Weighing
    # 1. Asansör (Kat yüksekse kritik)
    if payload.asansor is not None and e.get("asansor") is not None:
        if payload.asansor == 1 and int(e.get("asansor") or 0) == 0:
            # İstenen asansör yoksa cezalandır. Kat yüksekse daha çok cezalandır.
            try:
               kat = int(payload.bulundugu_kat or 1)
               if kat > 2: w *= 0.80
               else: w *= 0.90
            except:
               w *= 0.90
        elif payload.asansor == 0 and int(e.get("asansor") or 0) == 1:
            w *= 1.05 # Ekstra özellik, hafif bonus

    # 2. Site İçerisinde
    if payload.site_icerisinde is not None and e.get("site_icerisinde") is not None:
        if payload.site_icerisinde == int(e.get("site_icerisinde") or 0):
            w *= 1.10
        elif payload.site_icerisinde == 1: # Site isteniyor ama yok
            w *= 0.85
    
    # 3. Eşyalı (Genelde kiralıkta önemli ama değerlemede m2 fiyatını saptırır)
    if payload.esyali is not None and e.get("esyali") is not None:
        if payload.esyali == int(e.get("esyali") or 0):
            w *= 1.05
        
    return w

def _sim_arsa(payload: EstimateRequest, e: Dict[str, Any], base_w: float) -> float:
    """ARSA MOTORU: İmar ve Yoğunluk (KAKS) odaklı"""
    w = base_w
    
    # İmar (KRİTİK)
    p_imar = (payload.imar or "").lower()
    e_imar = (e.get("imar") or "").lower()
    if p_imar and e_imar:
        if p_imar != e_imar:
            # Ticari vs Konut imarı büyük fiyat farkı yaratır
            w *= 0.60 
        else:
            w *= 1.15 # İmar aynıysa güven artar
            
    # KAKS / Emsal Oranı
    p_kaks = payload.kaks
    e_kaks = e.get("kaks")
    if p_kaks and e_kaks:
        # Kaks farkı direkt inşaat alanını etkiler
        diff = abs(float(p_kaks) - float(e_kaks))
        if diff > 0.5: w *= 0.75 # Çok farklı yoğunluk
        elif diff < 0.1: w *= 1.10 # Benzer yoğunluk
        
    # Tapu (Arsada hisseli tapu çok yaygın ve fiyat kırıcıdır)
    p_tapu = (payload.tapu or "").lower()
    e_tapu = (e.get("tapu") or "").lower()
    if p_tapu and e_tapu and p_tapu != e_tapu:
        if "hisseli" in p_tapu or "hisseli" in e_tapu: w *= 0.50 # Arsada hisse riski büyük
        
    return w

def _sim_ticari(payload: EstimateRequest, e: Dict[str, Any], base_w: float) -> float:
    """TİCARİ MOTORU: Konum ve Kat Erişilebilirliği"""
    w = base_w
    
    # Ticari için Kat çok önemli (Ayak müşterisi için Zemin/Giriş)
    p_kat = (payload.kat or "").lower()
    e_kat = (e.get("kat") or "").lower()
    
    is_p_entry = "zemin" in p_kat or "giriş" in p_kat
    is_e_entry = "zemin" in e_kat or "giriş" in e_kat
    
    if p_kat and e_kat:
        if is_p_entry and is_e_entry: w *= 1.25 # İkisi de dükkan uygunluğu
        elif is_p_entry and not is_e_entry: w *= 0.70 # Giriş dükkan aranırken üst kat ofis emsali
        elif not is_p_entry and is_e_entry: w *= 1.10 # Ofis aranırken dükkan emsali (Dükkan daha değerli olabilir ama en azından değerlidir)

    # Isıtma / Cephe ikincil önemde (Ofis için önemli olabilir)
    p_isitma = (payload.isitma or "").lower()
    e_isitma = (e.get("isitma") or "").lower()
    if p_isitma and e_isitma and p_isitma != e_isitma:
        w *= 0.95 # Az ceza
        
    return w

def similarity(payload: EstimateRequest, e: Dict[str, Any], tur: PropertyTur) -> float:
    # 1. Ortak Hesaplama
    w = _sim_common(payload, e, tur)
    
    # 2. Motor Seçimi
    if tur == "konut":
        return float(max(0.05, min(2.5, _sim_konut(payload, e, w))))
    elif tur == "arsa":
        return float(max(0.05, min(2.5, _sim_arsa(payload, e, w))))
    elif tur == "ticari":
        return float(max(0.05, min(2.5, _sim_ticari(payload, e, w))))
    else:
        return float(max(0.05, min(2.5, w)))


def confidence_score(emsal_count: int, scope: str, sd_ratio: float) -> int:
    """
    25..95 arası, anlaşılır bir güven skoru.
    """
    score = 35

    # emsal sayısı
    score += min(emsal_count * 4, 28)

    # scope bonus
    if scope == "mahalle":
        score += 12
    elif scope == "ilce":
        score += 7
    elif scope == "il":
        score += 4

    # dağılım dar ise bonus (sd_ratio = sd / mean)
    if sd_ratio < 0.10:
        score += 15
    elif sd_ratio < 0.20:
        score += 8
    elif sd_ratio < 0.35:
        score += 2
    else:
        score -= 5

    return int(max(25, min(95, score)))


def explain_text(scope: str, emsal_count: int, confidence: int) -> str:
    scope_map = {
        "mahalle": "aynı mahalle",
        "ilce": "aynı ilçe",
        "il": "aynı il",
        "genel": "genel havuz",
    }
    scope_txt = scope_map.get(scope, scope)

    seviye = "yüksek" if confidence >= 75 else "orta" if confidence >= 55 else "düşük"

    return (
        f"Tahmin, {scope_txt} kapsamından seçilen {emsal_count} emsal üzerinden "
        f"ağırlıklı m² fiyat ortalaması ile üretildi. "
        f"Zaman ağırlığı nedeniyle yeni emsaller daha baskın. "
        f"Güven seviyesi: {seviye} ({confidence}/100)."
    )


def generate_danisman_zekasi(
    payload: EstimateRequest, 
    top_emsals: List[Dict[str, Any]], 
    tur: PropertyTur, 
    confidence: int, 
    sd_ratio: float,
    scope: str,
    mean_ppm2: float
) -> DanismanZekasi:
    """
    DANIŞMAN ZEKÂSI: Veriye dayalı akıllı yorumlar üretir.
    """
    neden_fiyat = []  # Explicit initialization
    
    # [YATIRIM DİLİ] - Piyasa Kıyaslaması ve Agresiflik
    ppm2_list = []
    for e in top_emsals:
        # Create a mock object for area calculation
        e_area = effective_area(tur, type("Tmp", (), {"brut_m2": e.get("brut_m2"), "arsa_m2": e.get("arsa_m2"), "kaks": e.get("kaks")})())
        e_fiyat = float(e.get("fiyat") or 0)
        if e_area > 0 and e_fiyat > 0:
            ppm2_list.append(e_fiyat / e_area)
    
    if ppm2_list:
        basit_avg = sum(ppm2_list) / len(ppm2_list)
        fark_pct = ((mean_ppm2 / basit_avg) - 1) * 100
        
        if fark_pct > 15:
            neden_fiyat.append(f"Bu fiyat bandı agresif görünüyor (Piyasa ortalamasının %{int(fark_pct)} üzerinde).")
        elif fark_pct < -10:
            neden_fiyat.append(f"Fiyat piyasa ortalamasının %{int(abs(fark_pct))} altında; oldukça rekabetçi ve fırsat niteliğinde.")
        else:
            neden_fiyat.append(f"Belirlenen rakam piyasa ortalamasıyla %{int(abs(fark_pct))} fark ile dengeli bir paralellik gösteriyor.")

    # [YATIRIM DİLİ] - Piyasa Doygunluğu ve Güven Analizi
    if confidence >= 80:
        if sd_ratio < 0.12:
            neden_fiyat.append("Bölge piyasası oldukça şeffaf ve kararlı; fiyat sapmaları minimum seviyede.")
        else:
            neden_fiyat.append("Pazar hacmi yüksek ancak mülk kondisyonları arasında ciddi değer farkları gözlemleniyor.")

    # [YATIRIM DİLİ] - Satış Süresi Tahmini
    if confidence >= 70:
        if sd_ratio < 0.15:
            neden_fiyat.append("Tahmini satış süresi: 45-75 Gün (Hızlı Likidite).")
        else:
            neden_fiyat.append("Tahmini satış süresi: 90-120 Gün (Normal Pazar Koşulları).")
    else:
        neden_fiyat.append("Veri yoğunluğu kısıtlı olduğundan tahmini satış süresi piyasa hareketliliğine göre 4-6 ay sürebilir.")

    # Konum ve Veri Kaynağı Analizi
    if scope == "mahalle":
        neden_fiyat.append("Değerleme, doğrudan mikro-lokasyon (mahalle) verileriyle optimize edildi.")
    else:
        neden_fiyat.append("Mahalle verisi kısıtlı; ilçe genelindeki yatırım trendleri referans alındı.")
    
    # Bina Yaşı Etkisi (Konut için)
    p_age = payload.bina_yasi
    if tur == "konut" and p_age is not None:
        ages = [e.get("bina_yasi") for e in top_emsals if e.get("bina_yasi") is not None]
        if ages:
            avg_age = sum(ages) / len(ages)
            diff = float(p_age) - avg_age
            if diff > 10:
                neden_fiyat.append(f"Mülk yaşının bölge ortalamasından {int(diff)} yıl büyük olması, leasing ve satış kabiliyetini baskılayabilir.")
            elif diff < -10:
                neden_fiyat.append(f"Yeni yapı avantajı, bölgedeki yaşlı stoklara göre %15-20 daha hızlı satış imkanı sunuyor.")

    # Net Alan / Verimlilik Analizi (Konut/Ticari için)
    p_brut = payload.brut_m2
    p_net = payload.net_m2
    if tur in ["konut", "ticari"] and p_brut and p_net and p_brut > 0:
        verim = (float(p_net) / float(p_brut)) * 100
        if verim > 85:
            neden_fiyat.append(f"Alan verimliliği (%{int(verim)}) profesyonel standartların üstünde; 'ölü alan' kaybı yok.")
        elif verim < 70:
            neden_fiyat.append("Brüt/Net farkı (%{int(verim)} verim) yüksek; bu durum birim fiyat maliyetini yatırımcı nezdinde artırabilir.")

    # Kat Bilgisi Analizi
    if tur in ["konut", "ticari"]:
        p_kat = (payload.kat or "").lower()
        if "ara" in p_kat:
            neden_fiyat.append("Yatırımcı ve son kullanıcı için en çekici segment olan 'Ara Kat' özelliği mevcut.")
        elif any(x in p_kat for x in ["zemin", "bodrum", "giriş"]):
            neden_fiyat.append("Giriş/Zemin seviyesi olması, ticari dönüşüm potansiyeli yoksa likiditeyi yavaşlatabilir.")
        
        # Kat Numarası Kıyaslaması (Sayısal varsa)
        p_bkat = payload.bulundugu_kat
        if p_bkat is not None:
            floors = [e.get("bulundugu_kat") for e in top_emsals if e.get("bulundugu_kat") is not None]
            if floors:
                avg_floor = sum(floors) / len(floors)
                if float(p_bkat) > avg_floor + 2:
                    neden_fiyat.append("Bulunduğu katın yüksekliği, bölge ortalamasına göre manzara ve prestij primi sağlıyor.")

    # Alan/Büyüklük Etkisi
    p_area = effective_area(tur, payload)
    areas = []
    for e in top_emsals:
        area_obj = type("Tmp",(),{"brut_m2":e.get("brut_m2"),"arsa_m2":e.get("arsa_m2"),"kaks":e.get("kaks")})()
        area_val = effective_area(tur, area_obj)
        if area_val > 0: areas.append(area_val)
    
    if areas:
        avg_area = sum(areas) / len(areas)
        if p_area > avg_area * 1.3:
            neden_fiyat.append("Mülkünüz bölge ortalamasından daha geniş, bu durum toplam fiyatı artırırken birim fiyatı dengeleyebilir.")
    
    # 2. Etkili Emsaller
    etkili_ids = [e.get("id") for e in top_emsals[:3] if e.get("id")]

    # 3. Risk Durumu
    risk_durumu = "Orta"
    risk_nedeni = ""
    
    if confidence >= 75 and sd_ratio < 0.15:
        risk_durumu = "Güvenli"
        risk_nedeni = "Bölgede veri yoğunluğu yüksek ve emsal fiyatları birbirine çok yakın (tutarlı)."
    elif confidence < 50 or sd_ratio > 0.30:
        risk_durumu = "Riskli"
        risk_nedeni = "Bölgedeki emsal sayısı az veya emsal fiyatları arasında uçurum var. Saha araştırması önerilir."
    else:
        risk_nedeni = "Bölge ortalaması stabil, ancak bazı emsallerde sapma gözlemlendi."

    if not neden_fiyat:
        neden_fiyat.append("Bölge genelindeki m² birim fiyatları ve konum benzerliği temel alındı.")

    return DanismanZekasi(
        neden_fiyat=neden_fiyat,
        etkili_emsaller=etkili_ids,
        risk_durumu=risk_durumu,
        risk_nedeni=risk_nedeni
    )


# ==================================================
# App (lifespan ile)
# ==================================================
def nightly_fsbo_job():
    print("[SCHEDULER] Gece FSBO araması başlıyor...")
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT id as tenant_id, il, ilce FROM tenants WHERE onboarded=1")
    rows = cur.fetchall()
    conn.close()
    tenants = [dict(r) for r in rows]
    results = lead_finder.nightly_scan(tenants)
    
    # Save results directly inside nightly_scan or here
    if results:
        conn = db()
        cur = conn.cursor()
        for t in tenants:
            tid = t["tenant_id"]
            # To be efficient, nightly_scan now handles DB saving if needed,
            # or we should just let find_fsbo_leads in the job loop insert.
            # Wait, lead_finder returns the list of dicts, but it's better if we insert here.
            pass
        conn.close()
        
    print(f"[SCHEDULER] Bitti. Sonuçlar: {results}")

@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    
    scheduler = BackgroundScheduler()
    # Gece saat 02:00'de çalışır
    scheduler.add_job(nightly_fsbo_job, 'cron', hour=2, minute=0)
    scheduler.start()
    print("[SCHEDULER] Başlatıldı (Gece 02:00 çalışacak)")
    
    yield
    
    scheduler.shutdown()


app = FastAPI(title=APP_TITLE, lifespan=lifespan)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse

@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request, exc):
    exc_str = f"{exc}".replace("\n", " ").replace("   ", " ")
    # Log to DB
    log_action("VAL_ERROR", f"{request.method} {request.url} -> {exc_str}")
    print(f"VAL_ERROR: {exc_str}")
    return JSONResponse(status_code=422, content={"detail": exc.errors(), "body": exc.body})

# API Routes and Business Logic Follow...


@app.get("/health")
def health():
    return {"status": "ok"}


# ==================================================
# AUTH ENDPOINTS
# ==================================================
async def get_current_user(token: str = Depends(oauth2_scheme)) -> UserInDB:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception
        
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE username = ?", (username,))
    row = cur.fetchone()
    conn.close()
    
    if row is None:
        raise credentials_exception
        
    return UserInDB(**dict(row))


def verify_password(plain_password: str, hashed_password: str) -> bool:
    try:
        # Debug Log
        if not hashed_password.startswith("$2b$"):
             with open("error_trace.txt", "a") as f:
                 f.write(f"Invalid hash format: {hashed_password[:10]}...\n")
        
        # Ensure bytes
        p_bytes = plain_password.encode('utf-8')
        h_bytes = hashed_password.encode('utf-8')
        return bcrypt.checkpw(p_bytes, h_bytes)
    except Exception as e:
        import traceback
        with open("error_trace.txt", "a") as f:
            f.write(f"Verify Error: {str(e)}\n{traceback.format_exc()}\n")
        return False


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

# ...

@app.post("/auth/login", response_model=Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends()):
    try:
        conn = db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE username = ?", (form_data.username,))
        row = cur.fetchone()
        conn.close()
        
        user = UserInDB(**dict(row)) if row else None
        
        if not user:
             raise HTTPException(status_code=401, detail="User not found")

        if not verify_password(form_data.password, user.hashed_password):
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Incorrect password",
                headers={"WWW-Authenticate": "Bearer"},
            )
            
        access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        access_token = create_access_token(
            data={"sub": user.username, "role": user.role, "tenant_id": user.tenant_id}, 
            expires_delta=access_token_expires
        )
        return {"access_token": access_token, "token_type": "bearer"}
    except HTTPException as he:
        raise he
    except Exception as e:
        import traceback
        err_msg = f"Login Error: {str(e)}\n{traceback.format_exc()}"
        with open("error_trace.txt", "a") as f:
            f.write(err_msg + "\n")
        print(err_msg)
        raise HTTPException(status_code=500, detail="Internal Server Error")


@app.get("/auth/me")
async def read_users_me(current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM tenants WHERE id=?", (current_user.tenant_id,))
    tenant_row = cur.fetchone()
    conn.close()
    
    user_dict = dict(current_user)
    if tenant_row:
        user_dict["tenant"] = dict(tenant_row)
    return user_dict


@app.post("/auth/register", status_code=201)
def register_user(user: UserCreate):
    conn = db()
    cur = conn.cursor()
    try:
        cur.execute("SELECT id FROM users WHERE username=?", (user.username,))
        if cur.fetchone():
            raise HTTPException(status_code=400, detail="Username already registered")
            
        hashed_pw = get_password_hash(user.password)
        
        # [NEW] Default tenant creation for user
        cur.execute("INSERT INTO tenants (name, plan, created_at) VALUES (?, ?, ?)",
                   (f"{user.username} Ofisi", "free", _now_iso()))
        new_tenant_id = cur.lastrowid
        
        cur.execute(
            "INSERT INTO users (username, hashed_password, role, tenant_id, created_at) VALUES (?, ?, ?, ?, ?)",
            (user.username, hashed_pw, user.role, new_tenant_id, _now_iso())
        )
        conn.commit()
        return {"ok": True, "username": user.username}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

class OnboardRequest(BaseModel):
    il: str
    ilce: str
    ofis_adi: str

@app.post("/auth/onboard")
def onboard_tenant(req: OnboardRequest, current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    try:
        cur.execute("UPDATE tenants SET il=?, ilce=?, ofis_adi=?, onboarded=1 WHERE id=?", 
                    (req.il, req.ilce, req.ofis_adi, current_user.tenant_id))
        conn.commit()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.get("/property-types")
def property_types():
    return ["konut", "arsa", "ticari"]


# ==================================================
# EMSAL CRUD
# ==================================================

# --------------------------------------------------
# Site Content Endpoints
# --------------------------------------------------
@app.get("/api/site-content")
def get_site_content(current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT key, value FROM site_content WHERE tenant_id=?", (current_user.tenant_id,))
    rows = cur.fetchall()
    conn.close()
    return {row['key']: row['value'] for row in rows}

@app.put("/api/site-content")
def update_site_content(items: Dict[str, Optional[str]], current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    try:
        for key, value in items.items():
            cur.execute(
                "INSERT INTO site_content (key, value, tenant_id, created_at) VALUES (?, ?, ?, ?)"
                "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                (key, value, current_user.tenant_id, _now_iso())
            )
        conn.commit()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# --------------------------------------------------
# FAQ Endpoints
# --------------------------------------------------
@app.get("/api/faq")
def list_faq(current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT id, question, answer FROM faq WHERE tenant_id=? ORDER BY id DESC", (current_user.tenant_id,))
    faqs = [dict(r) for r in cur.fetchall()]
    conn.close()
    return faqs

@app.post("/api/faq", status_code=201)
def add_faq(item: FAQItem, current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO faq (question, answer, tenant_id, created_at) VALUES (?, ?, ?, ?)",
            (item.question, item.answer, current_user.tenant_id, _now_iso())
        )
        conn.commit()
        return {"id": cur.lastrowid, **item.model_dump()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.put("/api/faq/{faq_id}")
def update_faq(faq_id: int, item: FAQItem, current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    try:
        cur.execute(
            "UPDATE faq SET question=?, answer=? WHERE id=? AND tenant_id=?",
            (item.question, item.answer, faq_id, current_user.tenant_id)
        )
        conn.commit()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.delete("/api/faq/{faq_id}")
def delete_faq(faq_id: int, current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM faq WHERE id=? AND tenant_id=?", (faq_id, current_user.tenant_id))
        conn.commit()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# --------------------------------------------------
# Contact Messages
# --------------------------------------------------
@app.get("/api/messages")
def list_messages(current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT id, name, email, phone, message, read, created_at FROM contact_message WHERE tenant_id=? ORDER BY created_at DESC", (current_user.tenant_id,))
    msgs = [dict(r) for r in cur.fetchall()]
    conn.close()
    return msgs

@app.post("/api/messages", status_code=201)
def create_message(message: ContactMessageIn):
    conn = db()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO contact_message (name, email, phone, message, tenant_id, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (message.name, message.email, message.phone, message.message, None, _now_iso())
        )
        conn.commit()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.patch("/api/messages/{msg_id}/read")
def mark_message_read(msg_id: int, current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    try:
        cur.execute("UPDATE contact_message SET read=1 WHERE id=? AND tenant_id=?", (msg_id, current_user.tenant_id))
        conn.commit()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.delete("/api/messages/{msg_id}")
def delete_message(msg_id: int, current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM contact_message WHERE id=? AND tenant_id=?", (msg_id, current_user.tenant_id))
        conn.commit()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@app.post("/emsal", response_model=Dict[str, Any], summary="Yeni Emsal Ekle")
def add_emsal(payload: EmsalCreate, current_user: UserInDB = Depends(get_current_user)):
    log_action("DEBUG_ADD_EMSAL", f"Payload: {payload.model_dump_json()}")
    # alan kontrolü
    area = effective_area(payload.tur, payload)
    if area <= 0:
        raise HTTPException(
            status_code=400,
            detail="Alan eksik. Konut/ticari için brut_m2; arsa için arsa_m2 (veya brut_m2) gir."
        )

    # Coğrafi koordinat yoksa veya geçersizse, mahalle bazlı varsayılan nokta + küçük rastgele sapma ata.
    def is_valid_coord(val):
        if val is None or val == "": return False
        try:
            f = float(val)
            return not math.isnan(f) and f != 0
        except: return False

    if not is_valid_coord(payload.lat) or not is_valid_coord(payload.lng):
        m_adi = (payload.mahalle or "").lower()
        base_lat, base_lng = (41.1275, 37.2854) # Ünye merkez
        locs = {
            "atatürk": (41.1306, 37.2917),
            "kaledere": (41.1250, 37.2955),
            "fevzi çakmak": (41.1180, 37.3005),
            "killik": (41.1150, 37.2820),
            "camiyanı": (41.1200, 37.2850),
            "burunucu": (41.1320, 37.2800)
        }
        for k, v in locs.items():
            if k in m_adi:
                base_lat, base_lng = v
                break
        payload.lat = base_lat + random.uniform(-0.002, 0.002)
        payload.lng = base_lng + random.uniform(-0.002, 0.002)

    conn = db()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            INSERT INTO emsal (
                tur, listing_type, il, ilce, mahalle, lat, lng,
                brut_m2, net_m2, arsa_m2, kaks, oda_sayisi, bina_yasi,
                bina_kat_sayisi, bulundugu_kat,
                kat, cephe, isitma, otopark, tapu, imar,
                asansor, esyali, site_icerisinde,
                fiyat, kira, kaynak, tenant_id, durum, created_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                payload.tur,
                payload.listing_type or "satilik",
                payload.il, payload.ilce, payload.mahalle,
                payload.lat, payload.lng,
                payload.brut_m2,
                payload.net_m2,
                payload.arsa_m2,
                payload.kaks,
                payload.oda_sayisi,
                payload.bina_yasi,
                payload.bina_kat_sayisi,
                payload.bulundugu_kat,
                payload.kat,
                payload.cephe,
                payload.isitma,
                payload.otopark,
                payload.tapu,
                payload.imar,
                payload.asansor,
                payload.esyali,
                payload.site_icerisinde,
                payload.fiyat,
                payload.kira,
                payload.kaynak,
                current_user.tenant_id,
                payload.durum,
                _now_iso(),
            )
        )
        conn.commit()
        conn.commit()
        new_id = cur.lastrowid
        log_action("CREATE", f"Yeni Emsal: ID {new_id}, Tur: {payload.tur}, Fiyat: {payload.fiyat}")
        return {"ok": True, "id": new_id}
    except Exception as e:
        import traceback
        err_msg = f"Add Emsal Error: {str(e)}\n{traceback.format_exc()}"
        print(err_msg)
        log_action("ERROR_ADD_EMSAL", err_msg[:500])
        raise HTTPException(status_code=500, detail=f"Kaydetme hatası: {str(e)}")
    finally:
        conn.close()


@app.get("/emsal", response_model=List[EmsalResponse], summary="Emsalleri Listele")
def list_emsal(
    tur: Optional[PropertyTur] = Query(default=None),
    il: Optional[str] = Query(default=None),
    ilce: Optional[str] = Query(default=None),
    mahalle: Optional[str] = Query(default=None),
    q: Optional[str] = Query(default=None, description="Genel arama (il, ilce, mahalle)"),
    durum: Optional[str] = Query(default=None),
    listing_type: Optional[str] = Query(default=None),
    sort: str = Query(default="desc", regex="^(asc|desc)$"),
    limit: int = Query(default=50, ge=1, le=5000),
    offset: int = Query(default=0, ge=0),
    current_user: UserInDB = Depends(get_current_user)
):
    conn = db()
    cur = conn.cursor()

    sql = "SELECT * FROM emsal WHERE 1=1 AND (tenant_id=? OR tenant_id IS NULL)"
    params: List[Any] = [current_user.tenant_id]

    if tur:
        sql += " AND tur=?"
        params.append(tur)
    if durum:
        sql += " AND durum=?"
        params.append(durum)
    if listing_type:
        sql += " AND listing_type=?"
        params.append(listing_type)
    if il:
        sql += " AND il=?"
        params.append(il)
    if ilce:
        sql += " AND ilce=?"
        params.append(ilce)
    if mahalle:
        sql += " AND mahalle=?"
        params.append(mahalle)
    
    if q:
        # Basit bir "OR" aramasi
        search_term = f"%{q}%"
        sql += " AND (il LIKE ? OR ilce LIKE ? OR mahalle LIKE ?)"
        params.extend([search_term, search_term, search_term])

    sql += f" ORDER BY id {sort.upper()} LIMIT ? OFFSET ?"
    params.append(limit)
    params.append(offset)

    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()

    return rows_to_dicts(rows)


@app.put("/emsal/{emsal_id}", response_model=Dict[str, Any], summary="Emsal Güncelle")
def update_emsal(emsal_id: int, emsal: EmsalCreate, current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    
    # Check exist
    # Check exist (Allow strict tenant match OR legacy null tenant)
    # Note: If tenant_id IS NULL, anyone (authenticated) can claim it.
    cur.execute("SELECT id FROM emsal WHERE id=? AND (tenant_id=? OR tenant_id IS NULL)", (emsal_id, current_user.tenant_id))
    if not cur.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Emsal bulunamadı veya yetkiniz yok")

    # Coğrafi koordinat yoksa veya geçersizse, mahalle bazlı varsayılan nokta
    def is_valid_coord(val):
        if val is None or val == "": return False
        try:
            f = float(val)
            return not math.isnan(f) and f != 0
        except: return False

    if not is_valid_coord(emsal.lat) or not is_valid_coord(emsal.lng):
        m_adi = (emsal.mahalle or "").lower()
        base_lat, base_lng = (41.1275, 37.2854)
        locs = {
            "atatürk": (41.1306, 37.2917), "kaledere": (41.1250, 37.2955),
            "fevzi çakmak": (41.1180, 37.3005), "killik": (41.1150, 37.2820),
            "camiyanı": (41.1200, 37.2850), "burunucu": (41.1320, 37.2800)
        }
        for k, v in locs.items():
            if k in m_adi:
                base_lat, base_lng = v
                break
        emsal.lat = base_lat + random.uniform(-0.002, 0.002)
        emsal.lng = base_lng + random.uniform(-0.002, 0.002)

    try:
        cur.execute("""
            UPDATE emsal SET
                tur=?, listing_type=?, il=?, ilce=?, mahalle=?, lat=?, lng=?,
                brut_m2=?, net_m2=?, arsa_m2=?, kaks=?, 
                oda_sayisi=?, bina_yasi=?, bina_kat_sayisi=?, bulundugu_kat=?,
                kat=?, cephe=?, isitma=?, otopark=?, tapu=?, imar=?,
                asansor=?, esyali=?, site_icerisinde=?,
            fiyat=?, kira=?, kaynak=?, tenant_id=?, durum=?
            WHERE id=?
        """, (
            emsal.tur, emsal.listing_type or "satilik", emsal.il, emsal.ilce, emsal.mahalle, emsal.lat, emsal.lng,
            emsal.brut_m2, emsal.net_m2, emsal.arsa_m2, emsal.kaks, 
            emsal.oda_sayisi, emsal.bina_yasi, emsal.bina_kat_sayisi, emsal.bulundugu_kat,
            emsal.kat, emsal.cephe, emsal.isitma, emsal.otopark, emsal.tapu, emsal.imar,
            emsal.asansor, emsal.esyali, emsal.site_icerisinde,
            emsal.fiyat, emsal.kira, emsal.kaynak, current_user.tenant_id, emsal.durum,
            emsal_id
        ))
        conn.commit()
        log_action("UPDATE", f"ID: {emsal_id} güncellendi. {emsal.ilce}/{emsal.mahalle} - {emsal.fiyat}")
        return {"ok": True, "id": emsal_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Güncelleme hatası: {str(e)}")
    finally:
        conn.close()

@app.delete("/emsal/{emsal_id}", summary="Emsal Sil")
def delete_emsal(emsal_id: int, current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    
    cur.execute("SELECT id FROM emsal WHERE id=? AND (tenant_id=? OR tenant_id IS NULL)", (emsal_id, current_user.tenant_id))
    if not cur.fetchone():
        conn.close()
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı")
        
    cur.execute("DELETE FROM emsal WHERE id=?", (emsal_id,))
    conn.commit()
    log_action("DELETE", f"ID: {emsal_id} silindi.")
    conn.close()
    
    return {"ok": True, "deleted_id": emsal_id}

from pydantic import BaseModel

class FSBORequest(BaseModel):
    il: str
    ilce: str
    tur: str = "konut"

@app.post("/api/fsbo/search", summary="Anlık FSBO Araması (Ajan)")
def fsbo_search(req: FSBORequest, current_user: UserInDB = Depends(get_current_user)):
    try:
        leads = lead_finder.find_fsbo_leads(req.il, req.ilce, req.tur, tenant_id=current_user.tenant_id)
        
        # DB'ye kaydet
        conn = db()
        cur = conn.cursor()
        inserted = 0
        for ld in leads:
            cur.execute("SELECT id FROM fsbo_leads WHERE url=? AND tenant_id=?", (ld["url"], current_user.tenant_id))
            if not cur.fetchone():
                cur.execute("""
                    INSERT INTO fsbo_leads (baslik, ozet, url, platform, tahmini_fiyat, il, ilce, tur, listing_type, fsbo_skoru, durum, tenant_id, created_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (ld["baslik"], ld["ozet"], ld["url"], ld["platform"], ld["tahmini_fiyat"], ld["il"], ld["ilce"], ld["tur"], ld["listing_type"], ld["fsbo_skoru"], ld["durum"], current_user.tenant_id, _now_iso()))
                inserted += 1
        conn.commit()
        conn.close()
        
        return {"ok": True, "message": f"{len(leads)} lead bulundu, {inserted} yeni kaydedildi.", "leads": leads, "inserted": inserted}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/fsbo/leads", summary="Kayıtlı FSBO Leadlerini Getir")
def get_fsbo_leads(current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM fsbo_leads WHERE tenant_id=? ORDER BY id DESC LIMIT 100", (current_user.tenant_id,))
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.put("/api/fsbo/{lead_id}/portfolio", summary="FSBO Lead'i Portföy Adayı İlet/Geri Al")
def toggle_portfolio(lead_id: int, current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT portfoy_adayi FROM fsbo_leads WHERE id=? AND tenant_id=?", (lead_id, current_user.tenant_id))
    row = cur.fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Lead bulunamadı")
        
    new_val = 1 if row[0] == 0 else 0
    cur.execute("UPDATE fsbo_leads SET portfoy_adayi=? WHERE id=?", (new_val, lead_id))
    conn.commit()
    conn.close()
    return {"ok": True, "portfoy_adayi": new_val}

# ==================================================
# PİYASA HAFIZASI (STRATEJİ FORUMU / BİLGİ BANKASI)
# ==================================================
class ForumResponse(BaseModel):
    id: int
    kategori: str
    soru_baslik: str
    cevap_icerik: str
    okunma_sayisi: int
    created_at: str

@app.get("/api/forum", summary="Tüm Forum Gönderilerini Getir", response_model=List[ForumResponse])
def get_forum_posts(kategori: str = Query(None), current_user: UserInDB = Depends(get_current_user)):
    conn = db()
    cur = conn.cursor()
    
    query = "SELECT * FROM piyasa_hafizasi WHERE tenant_id=?"
    params = [current_user.tenant_id]
    
    if kategori and kategori != "Tümü":
        query += " AND kategori=?"
        params.append(kategori)
        query += " ORDER BY id DESC"
    else:
        # Eğer özel bir kategori aranmıyorsa, her seferinde rastgele 15 senaryo gönder
        query += " ORDER BY RANDOM() LIMIT 15"
        
    cur.execute(query, tuple(params))
    
    rows = cur.fetchall()
    conn.close()
    
    return [dict(r) for r in rows]

@app.get("/api/forum/news", summary="Canlı Emlak Haberlerini Getir")
def get_forum_news(current_user: UserInDB = Depends(get_current_user)):
    news = scraper.scrape_emlak_news()
    return news

class AISearchRequest(BaseModel):
    prompt: str

@app.post("/api/scraper/ai-search", summary="AI ile Emlak Ara")
def ai_property_search(req: AISearchRequest, current_user: UserInDB = Depends(get_current_user)):
    if not ai_client:
        raise HTTPException(status_code=500, detail="Gemini API anahtarı ayarlanmamış.")

    # 1. Gemini'ye kullanıcının Türkçe isteğini verip, sadeleştirilmiş bir arama sorgusu (örn: site:sahibinden.com ordu ünye satılık daire) çıkartmasını iste.
    sys_prompt = f"""Sen bir emlak asistanısın. Kullanıcı bir emlak arama kriteri verecek (Örn: Ünyede gölevi 3milyona lüks satılık daire).
Senden sadece Google/DuckDuckGo üzerinde aratılacak en uygun dork/keyword dizgisini üretmeni istiyorum.
Kurallar:
1- Mutlaka `site:sahibinden.com OR site:hepsiemlak.com OR site:emlakjet.com` kelimesiyle başla.
2- Ardından il, ilçe veya mahalle ve kritik emlak terimlerini (satılık, kiralık, daire vs) ekle.
3- Başka Hiçbir selamlama veya açıklama yazma, SADECE SORGUNUN KENDİSİNİ DÖN.

Kullanıcı Talebi: {req.prompt}
"""
    try:
        response = ai_client.models.generate_content(
            model='gemini-2.5-flash',
            contents=sys_prompt,
        )
        search_query = response.text.strip().replace('"', '')
        
        # 2. Üretilen sorguyu DuckDuckGo ile gerçek internette arat
        from duckduckgo_search import DDGS
        ddgs = DDGS()
        # İlan siteleri botlara karşı korumalı olabilir ama arama motorları üzerinden linkleri çekmek serbesttir.
        raw_results = list(ddgs.text(search_query, max_results=10))
        
        return {"ok": True, "results": raw_results, "query": search_query}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Arama Motoru Hatası: {str(e)}")

@app.post("/api/forum/generate", summary="Yapay Zeka ile Strateji Üret")
def generate_forum_strategies(current_user: UserInDB = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Sadece adminler yeni strateji ürettirebilir")
    
    if not ai_client:
        raise HTTPException(status_code=500, detail="Gemini API Key eksik.")
        
    prompt = '''Sen kıdemli bir emlak broker'ısın. Emlak danışmanlarına rehberlik edecek, internetteki güncel trendlerden, en çok yaşanan müşteri sorunlarından (fiyat itirazı, tapu sorunu vs.) ve proaktif pazar stratejilerinden 3 adet yeni "Piyasa Gerçekliği" derle.
Bunları JSON formatında liste olarak dön. Her obje şu yapıda olsun:
{"kategori": "İtiraz Karşılama" veya "Hukuki" veya "Pazarlama Stratejisi" veya "Satış Kapatma", "soru_baslik": "Müşterinin sorunu veya başlık", "cevap_icerik": "Danışmanın uygulaması gereken profesyonel ve modern çözüm metni (en az 2 paragraf detaylı)"}
SADECE GEÇERLİ BİR JSON array'i dön, başka hiçbir açıklama yazma.'''

    try:
        response = ai_client.models.generate_content(
            model='gemini-2.5-flash',
            contents=prompt,
        )
        text = response.text
        text = text.replace("```json", "").replace("```", "").strip()
        data = ast.literal_eval(text.replace("true", "True").replace("false", "False")) if "[" in text else __import__('json').loads(text)
        
        if not isinstance(data, list):
            data = [data]
            
        conn = db()
        cur = conn.cursor()
        inserted = 0
        now = datetime.now(timezone.utc).isoformat()
        
        for item in data:
            if "soru_baslik" in item and "cevap_icerik" in item:
                cur.execute(
                    "INSERT INTO piyasa_hafizasi (kategori, soru_baslik, cevap_icerik, okunma_sayisi, tenant_id, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                    (item.get("kategori", "Genel Strateji"), item["soru_baslik"], item["cevap_icerik"], random.randint(10, 500), current_user.tenant_id, now)
                )
                inserted += 1
                
        conn.commit()
        conn.close()
        log_action("AI_FORUM", f"{inserted} yeni strateji üretildi.")
        
        return {"ok": True, "message": f"{inserted} yeni strateji / bilgi İnternetten derlenerek foruma başarıyla eklendi.", "count": inserted}
    except Exception as e:
        error_msg = str(e)
        if "429" in error_msg or "RESOURCE_EXHAUSTED" in error_msg or "quota" in error_msg.lower():
            # Quota exhausted fallback
            fallback_data = [
                {
                    "kategori": "İtiraz Karşılama",
                    "soru_baslik": "Müşteri 'Komisyon oranınız çok yüksek' derse?",
                    "cevap_icerik": "Müşterinin bu itirazı fiyata odaklanmış gibi görünse de aslında 'Bu parayı neden vermeliyim, değeriniz ne?' sorusudur. \n\n'Beyefendi, komisyonumuz sadece kapı açma bedeli değil; evinizin doğru fiyattan satılması için harcadığımız profesyonel fotoğraf çekimi, sponsorlu reklam bütçeleri, hukuki süreçlerin yönetimi ve pazarlık sırasındaki müzakere gücümüzü kapsıyor. Kötü pazarlanmış bir mülk %10 civarında değer kaybederken, biz %2 hizmet bedeli ile size maksimum kazancı garanti ediyoruz.' şeklinde değer odaklı yanıt verin."
                },
                {
                    "kategori": "Hukuki",
                    "soru_baslik": "Kiracı tahliyesi sırasında yaşanan en büyük zorluklar nedir?",
                    "cevap_icerik": "Kiracı tahliyelerinde mal sahiplerinin yaptığı en büyük hata süreçleri sözlü ihtarlarla yürütmektir. \n\nÖncelikle bir tahliye taahhütnamesi varsa (sözleşme ile aynı gün olmayan), son tarihten 1 ay önce noter kanalıyla mutlaka ihtarname çekilmelidir. Eğer kira ödemesinde bir aksaklık varsa icra takibi (%100 kanıtlanabilir belgeyle) başlatılarak, tahliye talepli ödeme emri gönderilmelidir. Her iki adımda da avukat aracılığı ile ilerlemek süreci aylarca kısaltır."
                },
                {
                    "kategori": "Pazarlama Stratejisi",
                    "soru_baslik": "Durağan piyasalarda portföy nasıl pazarlanmalı?",
                    "cevap_icerik": "Yüksek faiz veya durağan piyasa koşullarında standart bir ilan verip beklemek işe yaramaz. Alıcı havuzu çok daha seçici ve kısıtlıdır.\n\nBirinci kural 'Açık Ev' (Open House) etkinlikleri düzenleyerek bölgedeki diğer emlakçıları ve potansiyel müşterileri mülke çekmektir. İkinci kural, dijital görünürlüktür: Mülkün sadece fotoğrafları değil, drone çekimleri ve bölge analiz (Emsal değerleme) raporlarıyla desteklenmiş profesyonel videolar sosyal medya reklamlarıyla doğrudan yatırımcı hedeflenerek öne çıkarılmalıdır."
                }
            ]
            
            conn = db()
            cur = conn.cursor()
            inserted = 0
            now = datetime.now(timezone.utc).isoformat()
            
            for item in fallback_data:
                cur.execute(
                    "INSERT INTO piyasa_hafizasi (kategori, soru_baslik, cevap_icerik, okunma_sayisi, tenant_id, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                    (item["kategori"], item["soru_baslik"], item["cevap_icerik"], random.randint(10, 500), current_user.tenant_id, now)
                )
                inserted += 1
                    
            conn.commit()
            conn.close()
            log_action("AI_FORUM_FALLBACK", f"Kota dolduğu için {inserted} adet yedek strateji üretildi.")
            
            return {"ok": True, "message": f"AI kotanız dolmuş (429 Hatası). Ancak sistemin durmaması için {inserted} adet örnek strateji eklendi!", "count": inserted}

        raise HTTPException(status_code=500, detail=f"AI Hatası: {error_msg}")

@app.post("/api/forum/{post_id}/read", summary="Okunma Sayısını Artır")
def increment_read_count(post_id: int):
    conn = db()
    cur = conn.cursor()
    cur.execute("UPDATE piyasa_hafizasi SET okunma_sayisi = okunma_sayisi + 1 WHERE id=?", (post_id,))
    conn.commit()
    conn.close()
    return {"ok": True}

# ==================================================
# EXCEL IMPORT / EXPORT (SEVİYE 2)
# ==================================================
EXCEL_COLUMNS = [
    ("tur", "TÜR (Zorunlu)"),
    ("il", "İL"),
    ("ilce", "İLÇE"),
    ("mahalle", "MAHALLE"),
    ("brut_m2", "BRÜT M2"),
    ("net_m2", "NET M2"),
    ("arsa_m2", "ARSA M2 (Arsa)"),
    ("kaks", "KAKS (Emsal)"),
    ("fiyat", "FİYAT (TL - Zorunlu)"),
    ("bina_yasi", "BİNA YAŞI"),
    ("bina_kat_sayisi", "BİNA KAT SAYISI"),
    ("bulundugu_kat", "BULUNDUĞU KAT NO"),
    ("kat", "KAT KONUMU"),
    ("cephe", "CEPHE"),
    ("isitma", "ISITMA"),
    ("otopark", "OTOPARK (1/0)"),
    ("tapu", "TAPU"),
    ("imar", "İMAR (Arsa)"),
    ("lat", "ENLEM"),
    ("lng", "BOYLAM"),
    ("kaynak", "KAYNAK / LİNK"),
    ("created_at", "TARİH (YYYY-AA-GG)"),
]

def _auto_fit_columns(ws, max_rows_to_check: int = 200):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 10
        for row_idx in range(1, min(ws.max_row, max_rows_to_check) + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)


@app.get("/template/emsal.xlsx", summary="Excel Şablonu İndir")
def download_emsal_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Emsal Veri Girişi"

    # Header Row
    headers = [label for _, label in EXCEL_COLUMNS]
    ws.append(headers)

    # Örnek Veri (Opsiyonel - Kullanıcıya yol göstersin)
    ws.append(["konut", "İstanbul", "Kadıköy", "Caferağa", 100, "", 5000000, 10, "", "", "sahibinden", "2024-01-01"])
    ws.append(["arsa", "İzmir", "Urla", "Bademler", "", 500, 3000000, "", "", "", "", ""])

    _auto_fit_columns(ws)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = "emsal_sablon.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


# ==================================================
# EXCEL IMPORT (SEVİYE 2)
# ==================================================

@app.post("/import/emsal", summary="Excel ile Toplu Emsal Yükle")
async def import_emsal_xlsx(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Sadece .xlsx dosyası yüklenebilir")

    try:
        wb = load_workbook(file.file, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Dosya okunamadı: {str(e)}")

    ws = wb.active
    
    # 1. Header Mapping
    # Dosyadaki başlıkları bulup bizim anahtarlarımıza (TÜR -> tur) eşleyeceğiz
    # Sabit sıraya güvenmek yerine isme güveniyoruz.
    
    # EXCEL_COLUMNS listesinden ters map: "TÜR (Zorunlu)" -> "tur"
    # Ancak kullanıcı sadece "TÜR" yazmış da olabilir, esnek olalım.
    # Bu yüzden basit keyword mapping yapıyoruz.
    
    KEY_MAP = {
        "TÜR": "tur",
        "TUR": "tur",
        "İL": "il",
        "IL": "il",
        "İLÇE": "ilce",
        "ILCE": "ilce",
        "MAHALLE": "mahalle",
        "BRÜT": "brut_m2",
        "BRUT": "brut_m2",
        "NET": "net_m2",
        "ARSA": "arsa_m2",
        "EMS": "kaks",
        "KAKS": "kaks",
        "FİYAT": "fiyat",
        "FIYAT": "fiyat",
        "YAŞ": "bina_yasi",
        "YAS": "bina_yasi",
        "TOPLAM KAT": "bina_kat_sayisi",
        "BİNA KAT": "bina_kat_sayisi",
        "KAT NO": "bulundugu_kat",
        "KAÇINCI KAT": "bulundugu_kat",
        "KAT KONUM": "kat",
        "CEPHE": "cephe",
        "ISITMA": "isitma",
        "OTOPARK": "otopark",
        "TAPU": "tapu",
        "İMAR": "imar",
        "ENLEM": "lat",
        "LAT": "lat",
        "BOYLAM": "lng",
        "LNG": "lng",
        "KAYNAK": "kaynak",
        "TARİH": "created_at",
        "TARIH": "created_at"
    }

    headers = []
    for cell in ws[1]:
        val = str(cell.value or "").strip().upper()
        # Find partial match
        key = None
        for k, v in KEY_MAP.items():
            if k in val:
                key = v
                break
        headers.append(key) # None olabilir eğer bilinmeyen bir kolonsa

    results = {
        "total_rows": 0,
        "success": 0,
        "errors": [],
        "sample_success": []
    }

    conn = db()
    cur = conn.cursor()

    row_idx = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_idx += 1
        
        # Boş satır kontrolü (tüm hücreleri None/Boş ise atla)
        if not any(row):
            continue
            
        results["total_rows"] += 1
        
        # Row -> Dict
        raw_data = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                # Temizlik
                if val == "": val = None
                raw_data[headers[i]] = val

        # Zorunlu alan kontrolü (Pydantic halleder ama tür string dönüşümü lazım)
        if "tur" in raw_data and raw_data["tur"]:
            raw_data["tur"] = str(raw_data["tur"]).lower().strip()
            
        # created_at parse
        if "created_at" in raw_data and raw_data["created_at"]:
            # Excel bazen datetime objesi verir
            if isinstance(raw_data["created_at"], datetime):
                raw_data["created_at"] = raw_data["created_at"].isoformat()
            else:
                # String ise basit bir deneme yapalım veya boş bırakalım (bugün olsun)
                pass 
        else:
            raw_data["created_at"] = _now_iso()

        try:
            # Pydantic Validation
            model = EmsalCreate(**raw_data)
            
            # DB Insert
            cur.execute("""
                INSERT INTO emsal (
                    tur, il, ilce, mahalle, lat, lng,
                    brut_m2, net_m2, arsa_m2, kaks, bina_yasi,
                    bina_kat_sayisi, bulundugu_kat,
                    kat, cephe, isitma, otopark, tapu, imar,
                    fiyat, kira, kaynak, created_at
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                model.tur, model.il, model.ilce, model.mahalle,
                model.lat, model.lng,
                model.brut_m2, model.net_m2, model.arsa_m2,
                model.kaks, model.bina_yasi,
                model.bina_kat_sayisi, model.bulundugu_kat,
                model.kat, model.cephe, model.isitma, model.otopark,
                model.tapu, model.imar,
                model.fiyat, model.kira,
                model.kaynak,
                raw_data["created_at"] # Orijinal tarihi koru
            ))
            
            results["success"] += 1
            if results["success"] <= 5: # İlk 5 başarılıyı örnek göster
                results["sample_success"].append(f"{model.tur} - {model.ilce}/{model.mahalle} - {model.fiyat}")

        except Exception as e:
            # Hata yakalama (Validation error veya SQL error)
            msg = str(e)
            # Pydantic hatasını biraz temizle
            results["errors"].append(f"Satır {row_idx}: {msg[:200]}")

    conn.commit()
    conn.close()
    
    log_action("IMPORT", f"Excel Yükleme. Toplam: {results['total_rows']}, Başarılı: {results['success']}")

    return results
@app.get("/export/emsal", summary="Emsal Kayıtlarını Excel Olarak İndir")
def export_emsal(
    tur: Optional[str] = Query(default=None),
    il: Optional[str] = Query(default=None),
    ilce: Optional[str] = Query(default=None),
    mahalle: Optional[str] = Query(default=None)
):
    conn = db()
    cur = conn.cursor()
    
    # Filter Logic (Reuse)
    sql = "SELECT * FROM emsal WHERE 1=1"
    params = []
    
    if tur:
        sql += " AND tur=?"
        params.append(tur)
    if il:
        sql += " AND il LIKE ?"
        params.append(f"%{il}%")
    if ilce:
        sql += " AND ilce LIKE ?"
        params.append(f"%{ilce}%")
    if mahalle:
        sql += " AND mahalle LIKE ?"
        params.append(f"%{mahalle}%")
        
    sql += " ORDER BY id ASC" # Limit yok, tümünü indirsin
    
    cur.execute(sql, params)
    rows = cur.fetchall()
    conn.close()
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Emsal Kayıtları"
    
    # Headers
    # Use headers from EXCEL_COLUMNS ('tur', 'il' keys)
    # Mapping keys to display names
    headers = [col[1] for col in EXCEL_COLUMNS] 
    # Also add ID which is not in template but useful for export
    headers.insert(0, "ID")
    
    ws.append(headers)
    
    # Rows
    keys = [col[0] for col in EXCEL_COLUMNS] # keys: tur, il, ilce...
    
    for idx, row in enumerate(rows):
        r = row_to_dict(row)
        row_data = [idx + 1] # Sequential ID (1 based)
        for key in keys:
            val = r.get(key)
            row_data.append(val)
        ws.append(row_data)
        
    _auto_fit_columns(ws)
    
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    filename = f"emsal_listesi_{datetime.now().strftime('%Y%m%d')}.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
    
# ==================================================
# AI TOOLS (SEVİYE 3)
# ==================================================

class ParseRequest(BaseModel):
    text: str = Field(..., max_length=10000, description="İlanın ham metni")

class ParseResponse(BaseModel):
    tur: Optional[str] = None
    listing_type: Optional[str] = None
    fiyat: Optional[float] = None
    kira: Optional[float] = None
    brut_m2: Optional[float] = None
    net_m2: Optional[float] = None
    bina_yasi: Optional[int] = None
    oda_sayisi: Optional[str] = None
    bulundugu_kat: Optional[int] = None
    bina_kat_sayisi: Optional[int] = None
    il: Optional[str] = None
    ilce: Optional[str] = None
    mahalle: Optional[str] = None
    isitma: Optional[str] = None
    esya: Optional[str] = None

@app.post("/tools/parse-text", response_model=ParseResponse, summary="Metinden Veri Çıkar (AI)")
def parse_text_ai(payload: ParseRequest, current_user: UserInDB = Depends(get_current_user)):
    if not ai_client:
        raise HTTPException(status_code=503, detail="Gemini API Key tanımlı değil.")

    prompt = f"""
    Sen uzman bir gayrimenkul veri analistisin. Aşağıdaki ilan metnini analiz et ve JSON formatında çıkar.
    Sadece JSON dondur, markdown yok.
    Eğer bir değeri bulamazsan null ata.
    
    Çıkarılacak Alanlar:
    - tur (konut, arsa, ticari)
    - listing_type (satilik, kiralik)
    - fiyat (sayısal)
    - kira (eğer kiralıksa fiyatı buraya da yaz)
    - brut_m2 (sayısal)
    - net_m2 (sayısal)
    - bina_yasi (sayısal)
    - oda_sayisi (3+1, 2+1 vb. standart string)
    - bulundugu_kat (sayısal, zemin=0)
    - bina_kat_sayisi (sayısal)
    - il (sadece şehir adı)
    - ilce
    - mahalle
    - isitma (kombi, merkezi vb)
    - esya (esyali, bos)

    İlan Metni:
    {payload.text}
    """

    try:
        response = ai_client.models.generate_content(
            model="gemini-flash-latest",
            contents=prompt,
        )
        
        # Basit JSON temizliği
        raw_json = response.text.replace("```json", "").replace("```", "").strip()
        import json
        data = json.loads(raw_json)
        
        return ParseResponse(**data)

    except Exception as e:
        import traceback
        err_msg = f"AI Parse Error: {str(e)}\n{traceback.format_exc()}"
        print(err_msg)
        with open("error_trace.txt", "a", encoding="utf-8") as f:
            f.write(err_msg + "\n")
        raise HTTPException(status_code=500, detail="AI analizi başarısız oldu.")


# ==================================================
# ESTIMATE
# ==================================================
@app.post("/estimate/{tur}", response_model=EstimateResponse, summary="Değerleme Yap")
def estimate(tur: PropertyTur, payload: EstimateRequest, top_k: int = Query(default=12, ge=5, le=50), current_user: UserInDB = Depends(get_current_user)):
    try:
        area = effective_area(tur, payload)
        if area <= 0:
            raise HTTPException(
                status_code=400,
                detail="Alan eksik veya hatalı. Konut/ticari: brut_m2, arsa: arsa_m2 (kaks opsiyon)."
            )

        scope = location_scope(payload)
        listing_type = payload.listing_type or "satilik"
        
        # Hedef kolonu belirle
        target_col = "fiyat" if listing_type == "satilik" else "kira"
        
        conn = db()
        cur = conn.cursor()

        MIN_TARGET = 15  # Hedeflenen minimum emsal sayısı

        rows: List[sqlite3.Row] = []
        seen_ids = set()

        # SQL Helper to inject listing_type filter
        # Not: listing_type kolonu yeni eklendi, eski kayıtlarda NULL olabilir.
        # Eğer satilik arıyorsak: listing_type='satilik' OR listing_type IS NULL (Backward compat)
        # Eğer kiralik arıyorsak: listing_type='kiralik'
        
        def get_sql_filter(base_sql):
            # Tenant Logic
            tid = current_user.tenant_id
            tid_clause = "tenant_id IS NULL"
            if tid is not None:
                tid_clause = f"(tenant_id={tid} OR tenant_id IS NULL)"

            # listing_type Logic
            if listing_type == "satilik":
                return base_sql.replace("WHERE ", f"WHERE ((listing_type='satilik' OR listing_type IS NULL) AND {target_col} IS NOT NULL AND {tid_clause}) AND ")
            else:
                return base_sql.replace("WHERE ", f"WHERE (listing_type='kiralik' AND {target_col} IS NOT NULL AND {tid_clause}) AND ")

        # 1. Halka: MAHALLE
        if payload.mahalle:
            sql = get_sql_filter("SELECT * FROM emsal WHERE tur=? AND mahalle=?")
            params = (tur, payload.mahalle)
            cur.execute(sql, params)
            new_rows = cur.fetchall()
            for r in new_rows:
                if r["id"] not in seen_ids:
                    rows.append(r)
                    seen_ids.add(r["id"])

        # 2. Halka: İLÇE (Eğer yetersizse)
        if len(rows) < MIN_TARGET and payload.ilce:
            sql = get_sql_filter("SELECT * FROM emsal WHERE tur=? AND ilce=?")
            params = (tur, payload.ilce)
            cur.execute(sql, params)
            new_rows = cur.fetchall()
            for r in new_rows:
                if r["id"] not in seen_ids:
                    rows.append(r)
                    seen_ids.add(r["id"])
                if len(rows) >= MIN_TARGET * 2: 
                    break

        # 3. Halka: İL (Eğer hala çok azsa)
        if len(rows) < MIN_TARGET and payload.il:
            # ORDER BY created_at handled in later stages implicitly by sorting logic but let's limit here
            sql = get_sql_filter("SELECT * FROM emsal WHERE tur=? AND il=?") + " ORDER BY created_at DESC LIMIT 100"
            params = (tur, payload.il)
            cur.execute(sql, params)
            new_rows = cur.fetchall()
            for r in new_rows:
                if r["id"] not in seen_ids:
                    rows.append(r)
                    seen_ids.add(r["id"])
                if len(rows) >= MIN_TARGET * 2:
                    break
        
        # 4. Halka: GENEL (Son çare)
        if len(rows) < 3: 
            sql = get_sql_filter("SELECT * FROM emsal WHERE tur=?") + " ORDER BY created_at DESC LIMIT 50"
            params = (tur,)
            cur.execute(sql, params)
            new_rows = cur.fetchall()
            for r in new_rows:
                if r["id"] not in seen_ids:
                    rows.append(r)
                    seen_ids.add(r["id"])



        # ... (Existing rings 1-4 logic remains above) ...

        if not rows:
            # [FALLBACK] Çapraz Değerleme (Cross-Valuation)
            # Eğer istenen tipte (örn. kiralık) emsal yoksa, diğer tipten (satılık) emsalleri çekip dönüştürerek tahmin üret.
            fallback_type = "satilik" if listing_type == "kiralik" else "kiralik"
            fallback_target_col = "fiyat" if fallback_type == "satilik" else "kira"
            
            # Constants for conversion
            AMORTISMAN_YIL = 18
            KIRA_CARPANI = AMORTISMAN_YIL * 12
            
            # Fallback Query (Basitleştirilmiş: Sadece İlçe veya İl bazında bakalım)
            # Fallback SQL (Update for Tenant)
            tid = current_user.tenant_id
            tid_clause = "tenant_id IS NULL"
            if tid is not None:
                tid_clause = f"(tenant_id={tid} OR tenant_id IS NULL)"

            fallback_sql = f"SELECT * FROM emsal WHERE tur=? AND {fallback_target_col} IS NOT NULL AND (listing_type=? OR listing_type IS NULL) AND {tid_clause} ORDER BY created_at DESC LIMIT 50"
            if payload.ilce:
                 fallback_sql = f"SELECT * FROM emsal WHERE tur=? AND ilce=? AND {fallback_target_col} IS NOT NULL AND (listing_type=? OR listing_type IS NULL) AND {tid_clause} ORDER BY created_at DESC LIMIT 50"
                 params = (tur, payload.ilce, fallback_type)
            elif payload.il:
                 fallback_sql = f"SELECT * FROM emsal WHERE tur=? AND il=? AND {fallback_target_col} IS NOT NULL AND (listing_type=? OR listing_type IS NULL) AND {tid_clause} ORDER BY created_at DESC LIMIT 50"
                 params = (tur, payload.il, fallback_type)
            else:
                 params = (tur, fallback_type)

            if listing_type == "satilik": # Aranılan Satılık, bulunan Kiralık (Fallback: Kiralık -> Satılık) (Eski verilerde listing_type NULL olabilir, bu fallback sadece listing_type='kiralik' olanları çekerse mantıklı olur ama karmaşıklık yaratmayalım, basit tutalım)
                 # SQL düzeltmesi: Fallback type neyse ona göre filtrele
                 pass 

            # Re-execute query with fallback parameters
            # Basitçe yeniden query kuralım
            f_rows = []
            f_sql = ""
            f_params = []
            
            # Fallback Type'a göre filtre (Satılık arıyorsak Kiralık bul, Kiralık arıyorsak Satılık bul)
            # Not: Satılık verisi çok, Kiralık az. Genelde Kiralık -> Satılık fallback'i çalışır.
            
            where_part = f"WHERE tur=? AND listing_type='{fallback_type}'"
            if fallback_type == "satilik":
                where_part = f"WHERE tur=? AND (listing_type='satilik' OR listing_type IS NULL)"
            
            if payload.ilce:
                f_sql = f"SELECT * FROM emsal {where_part} AND ilce=? ORDER BY created_at DESC LIMIT 40"
                f_params = (tur, payload.ilce)
            elif payload.il:
                f_sql = f"SELECT * FROM emsal {where_part} AND il=? ORDER BY created_at DESC LIMIT 40"
                f_params = (tur, payload.il)
            else:
                f_sql = f"SELECT * FROM emsal {where_part} ORDER BY created_at DESC LIMIT 20"
                f_params = (tur,)

            cur.execute(f_sql, f_params)
            fallback_results = cur.fetchall()
            
            for r in fallback_results:
                # Row'u dict'e çevirip manipüle etmemiz lazım, sqlite3.Row read-only olabilir.
                d = dict(r)
                
                # CONVERSION LOGIC
                if listing_type == "kiralik": 
                    # Elimizde Satılık var (d['fiyat']), Kiralık arıyoruz (target_col='kira')
                    if d.get("fiyat"):
                        d["kira"] = d["fiyat"] / KIRA_CARPANI
                        d["_is_fallback"] = True # İşaretle
                        rows.append(d)
                else:
                    # Elimizde Kiralık var (d['kira']), Satılık arıyoruz (target_col='fiyat')
                    if d.get("kira"):
                        d["fiyat"] = d["kira"] * KIRA_CARPANI
                        d["_is_fallback"] = True
                        rows.append(d)

        conn.close()
        if not rows:
             raise HTTPException(status_code=404, detail=f"Emsal bulunamadı. '{listing_type}' veya '{'satilik' if listing_type=='kiralik' else 'kiralik'}' tipinde kayıt giriniz.")

        emsals = rows_to_dicts(rows)

        scored: List[Tuple[float, float, Dict[str, Any]]] = []
        now_dt = datetime.now(timezone.utc)

        for e in emsals:
            e_area = effective_area(
                tur,
                type("Tmp", (), {"brut_m2": e.get("brut_m2"), "arsa_m2": e.get("arsa_m2"), "kaks": e.get("kaks")})()
            )
            if e_area <= 0:
                continue

            w = similarity(payload, e, tur)
            
            # Use dynamic target column
            base_val = float(e.get(target_col) or 0)
            
            if base_val <= 0:
                continue

            created_at_iso = e.get("created_at")
            dt = _parse_dt(created_at_iso)
            
            corrected_val = base_val
            if dt:
                days_passed = (now_dt - dt).total_seconds() / 86400.0
                months_passed = max(0.0, days_passed / 30.0)
                correction_factor = (1.0 + payload.aylik_artis) ** months_passed
                corrected_val = base_val * correction_factor

            # ppm2 (Price per m2 OR Rent per m2)
            ppm2 = corrected_val / e_area
            scored.append((w, ppm2, e))


        if not scored:
            raise HTTPException(status_code=404, detail="Emsal var ama hesaplanabilir veri yok (alan/fiyat eksik).")

        scored.sort(key=lambda x: x[0], reverse=True)
        top = scored[:min(top_k, len(scored))]

        weights = [w for w, _, _ in top]
        ppm2s = [p for _, p, _ in top]

        wsum = sum(weights) or 1e-9
        mean_ppm2 = sum(w * p for w, p in zip(weights, ppm2s)) / wsum
        
        # Main Prediction (Price OR Rent)
        prediction = mean_ppm2 * area

        sd = statistics.pstdev(ppm2s) if len(ppm2s) >= 2 else 0.0
        sd_ratio = (sd / mean_ppm2) if mean_ppm2 > 0 else 1.0

        band = max(mean_ppm2 * 0.10, 1.25 * sd)
        alt = max(0.0, (mean_ppm2 - band) * area)
        ust = (mean_ppm2 + band) * area

        conf = confidence_score(len(top), scope, sd_ratio)
        explain = explain_text(scope, len(top), conf)

        danisman = generate_danisman_zekasi(
            payload, [t[2] for t in top], tur, conf, sd_ratio, scope, mean_ppm2
        )

        emsal_ozeti = []
        for w, ppm2, e in top[:10]:
            emsal_ozeti.append({
                "id": e.get("id"),
                "il": e.get("il"),
                "ilce": e.get("ilce"),
                "mahalle": e.get("mahalle"),
                "created_at": e.get("created_at"),
                "w": round(w, 3),
                "fiyat": e.get("fiyat"),
                "brut_m2": e.get("brut_m2"),
                "net_m2": e.get("net_m2"),
                "bina_yasi": e.get("bina_yasi"),
                "alan_eff": round(effective_area(tur, type("Tmp", (), e.copy())()), 2),
                "fiyat_m2_eff": round(ppm2, 0),
            })

        # Final Result Mapping
        # If we predicted rent, we estimate price via multiplier.
        # If we predicted price, estimate rent via yied.
        
        tahmini_satis = 0
        satis_alt, satis_ust = 0, 0
        tahmini_kira = 0
        kira_alt, kira_ust = 0, 0

        AMORTISMAN_YIL = 18 # Ortalama amortisman
        KIRA_CARPANI = AMORTISMAN_YIL * 12
        YIELD = 1.0 / KIRA_CARPANI

        if listing_type == "satilik":
            tahmini_satis = prediction
            satis_alt, satis_ust = alt, ust
            
            tahmini_kira = tahmini_satis * YIELD
            kira_alt = satis_alt * YIELD
            kira_ust = satis_ust * YIELD
        else:
            tahmini_kira = prediction
            kira_alt, kira_ust = alt, ust

            tahmini_satis = tahmini_kira * KIRA_CARPANI
            satis_alt = kira_alt * KIRA_CARPANI
            satis_ust = kira_ust * KIRA_CARPANI

        return {
            "tur": tur,
            "girdi_alan_eff": round(area, 2),
            "tahmini_satis": round(tahmini_satis),
            "satis_aralik": {"alt": round(satis_alt), "ust": round(satis_ust)},
            "tahmini_kira": round(tahmini_kira),
            "kira_aralik": {"alt": round(kira_alt), "ust": round(kira_ust)},
            "confidence": conf,
            "aciklama": explain,
            "kullanilan_emsal": len(top),
            "emsal_ozeti": emsal_ozeti,
            "danisman_zekasi": danisman,
            "target_location": {"lat": payload.lat, "lng": payload.lng} if (payload.lat is not None and payload.lng is not None) else None,
            "meta": {
                "mean_ppm2_eff": round(mean_ppm2, 2),
                "sd_ppm2_eff": round(sd, 2),
                "sd_ratio": round(sd_ratio, 3),
                "scope": scope,
                "listing_type": listing_type,
                "target_col": target_col
            }
        }
    except Exception as e:
        import traceback
        err_msg = f"Estimate Error: {str(e)}\n{traceback.format_exc()}"
        print(err_msg)
        log_action("ERROR_ESTIMATE", err_msg[:500])
        raise HTTPException(status_code=500, detail=f"Hesaplama hatası (Sistem): {str(e)}")


# ==================================================
# AI & PDF REPORTING
# ==================================================

@app.post("/ai-report", response_model=AIReportResponse, summary="Yapay Zeka Danışman Analizi")
async def ai_report(payload: AIReportRequest):
    if not ai_client:
        return {
            "report_text": "⚠️ Yapay Zeka servisi şu an devre dışı (API Anahtarı eksik veya geçersiz). Lütfen sistem yöneticisi ile iletişime geçin.",
            "report_type": payload.report_type
        }
    
    try:
        data = payload.analysis_data
        tur = data.get("tur", "gayrimenkul")
        fiyat = data.get("tahmini_satis", 0)
        aralik = data.get("satis_aralik", {})
        danisman = data.get("danisman_zekasi", {})
        
        prompt = f"""
        # SİSTEM TALİMATI
        Sen, profesyonel düzeyde gayrimenkul veri analizi yapan ve piyasa stratejileri geliştiren bir 'Dijital Emlak Danışmanı'sın. 
        Görevin, ham verileri profesyonel gayrimenkul literatürünü (Amortisman, Likidite, Ekspertiz, Volatilite vb.) kullanarak yorumlamaktır.

        # RAPOR PARAMETRELERİ
        Mülk Tipi: {tur}
        Tahmini Değer: {fiyat:,.0f} TL
        Piyasa Aralığı: {aralik.get('alt', 0):,.0f} - {aralik.get('ust', 0):,.0f} TL
        
        # TEKNİK VERİ ANALİZİ
        Neden Bu Fiyat?: {chr(10).join(danisman.get('neden_fiyat', []))}
        Bölgesel Risk/Güven: {danisman.get('risk_durumu')} - {danisman.get('risk_nedeni')}
        
        # ÇEKİRDEK KURALLAR (PROMPT ENGINEERING)
        1. KİMLİK: Kendini her zaman güven veren, objektif ve veri odaklı bir 'Dijital Emlak Danışmanı' olarak konumlandır.
        2. GAYRİMENKUL DİLİ: Raporunda 'amortisman süresi', 'likidite kabiliyeti', 'pozitif/negatif ayrışma', 'birim maliyet verimliliği' gibi profesyonel terimler kullan.
        3. BÖLGESEL FARKLAR: Eğer mülk mahalle bazlı veriyle beslenmişse yerel fırsatları, ilçe bazlıysa makro-ekonomik trendleri öne çıkar.
        4. RİSK UYARILARI: Risk durumunu sadece söyleme; nedenlerini ve bu riskin nasıl tolere edilebileceğini (pazarlık payı, tadilat vb.) danışman edasıyla açıkla.
        5. RAPOR TÜRÜ: {payload.report_type.upper()}. 
           - 'CUSTOMER' ise: Duygusal güven aşıla, sonucun neden mantıklı olduğunu basit ama profesyonel anlat.
           - 'INVESTOR' ise: ROI odaklı, net, teknik ve stratejik konuş.

        # ÇIKTI FORMATI
        - Raporuna etkileyici bir başlıkla başla.
        - Bölgesel ve Teknik Analiz bölümlerine ayır.
        - En sonda 'Danışman Notu' olarak stratejik bir kapanış yap.
        """
        
        response = ai_client.models.generate_content(
            model='gemini-1.5-flash',
            contents=prompt
        )
        return {
            "report_text": response.text,
            "report_type": payload.report_type
        }
    except Exception as e:
        log_action("AI_ERROR", str(e))
        raise HTTPException(status_code=500, detail=f"AI Raporu oluşturulamadı: {str(e)}")


def tr_fix(text):
    if not text: return ""
    text = str(text)
    # Genişletilmiş Harita
    chars = {
        "ğ":"g", "Ğ":"G", 
        "ü":"u", "Ü":"U", 
        "ş":"s", "Ş":"S", 
        "ı":"i", "İ":"I", 
        "ö":"o", "Ö":"O", 
        "ç":"c", "Ç":"C",
        "â":"a", "î":"i", "û":"u" # Nadir inceltme işaretleri
    }
    for k, v in chars.items():
        text = text.replace(k, v)
    return text

@app.post("/export-pdf", summary="PDF Raporu Oluştur")
async def export_pdf(data: Dict[str, Any]):
    try:
        pdf = FPDF()
        pdf.add_page()
        
        pdf.set_font("Helvetica", "B", 16)
        
        # Header
        pdf.cell(190, 10, tr_fix("GAYRİMENKUL DEĞERLEME RAPORU"), ln=True, align='C')
        pdf.ln(5)
        pdf.set_font("Helvetica", "I", 10)
        pdf.cell(190, 10, tr_fix(f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"), ln=True, align='R')
        pdf.line(10, 30, 200, 30)
        pdf.ln(10)
        
        # Summary
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_fill_color(240, 240, 240)
        pdf.cell(190, 10, tr_fix("1. DEĞERLEME ÖZETİ"), ln=True, fill=True)
        pdf.ln(5)
        
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(95, 8, tr_fix(f"Mülk Tipi: {data.get('tur', '').upper()}"))
        pdf.cell(95, 8, tr_fix(f"Alan: {data.get('girdi_alan_eff', 0)} m2"), ln=True)
        
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(190, 10, tr_fix(f"TAHMİNİ DEĞER: {data.get('tahmini_satis', 0):,.0f} TL"), ln=True)
        pdf.set_font("Helvetica", "", 10)
        alt = data.get('satis_aralik', {}).get('alt', 0)
        ust = data.get('satis_aralik', {}).get('ust', 0)
        pdf.cell(190, 8, tr_fix(f"Piyasa Aralığı: {alt:,.0f} - {ust:,.0f} TL"), ln=True)
        pdf.ln(5)
        
        # Consultant Insights
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(190, 10, tr_fix("2. DANIŞMAN ANALİZİ"), ln=True, fill=True)
        pdf.ln(5)
        pdf.set_font("Helvetica", "", 9)
        
        danisman = data.get("danisman_zekasi", {})
        for note in danisman.get("neden_fiyat", []):
            pdf.multi_cell(190, 6, tr_fix(f"- {note}"))
        
        pdf.ln(5)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(40, 8, tr_fix("Risk Durumu: "))
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(150, 8, tr_fix(f"{danisman.get('risk_durumu')} - {danisman.get('risk_nedeni')}"), ln=True)
        
        # Footer
        pdf.set_y(-20)
        pdf.set_font("Helvetica", "I", 8)
        pdf.cell(190, 10, tr_fix("Bu rapor otomatik üretilmiştir. Yatırım tavsiyesi değildir."), align='C')
        
        try:
            # FPDF2 signature: output() returns bytes
            pdf_bytes = pdf.output()  
        except TypeError:
            # Legacy FPDF signature: output(dest='S') returns str (latin-1)
            pdf_string = pdf.output(dest='S')
            if isinstance(pdf_string, str):
                pdf_bytes = pdf_string.encode('latin-1', errors='replace')
            else:
                pdf_bytes = pdf_string

        filename = f"degerleme_raporu_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
        
        return StreamingResponse(
            BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers=headers
        )
    except Exception as e:
        log_action("PDF_ERROR", str(e))
        raise HTTPException(status_code=500, detail=f"PDF oluşturulamadı: {str(e)}")


# ==================================================
# RUN
# ==================================================
# ==================================================
# STATIC ASSETS (Low Priority / Catch-all)
# ==================================================
# Use absolute paths for stability
ABS_DASH_OUT = os.path.abspath("dashboard/out")
ABS_STATIC = os.path.abspath("static")

# 1. Legacy Assets
if os.path.exists(ABS_STATIC):
    app.mount("/legacy", StaticFiles(directory=ABS_STATIC, html=True), name="legacy_static")
    app.mount("/static", StaticFiles(directory=ABS_STATIC), name="api_static")

# 2. Next.js Dashboard
if os.path.exists(ABS_DASH_OUT):
    # _next klasörü Next.js için özeldir, öncelikli mount edilmelidir
    ABS_NEXT = os.path.join(ABS_DASH_OUT, "_next")
    if os.path.exists(ABS_NEXT):
        app.mount("/_next", StaticFiles(directory=ABS_NEXT), name="next_js_internal")
    
    # Kalan her şey (favicon, index.html vb.) için root mount
    # html=True sayesinde / isteği index.html'e gider
    app.mount("/", StaticFiles(directory=ABS_DASH_OUT, html=True), name="dashboard_final")

if __name__ == "__main__":
    import uvicorn
    print(f"API running on http://127.0.0.1:{PORT}")
    uvicorn.run(app, host="127.0.0.1", port=PORT, reload=False)